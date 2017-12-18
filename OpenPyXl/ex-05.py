from openpyxl import load_workbook

# 載入 Excel 檔案（即：活頁簿），取得所有工作表的名稱
wb = load_workbook('../data/sample.xlsx')
print(wb.get_sheet_names())

# 依「工作表的名稱」取得工作表
sheet = wb.get_sheet_by_name('工作表1')
print(sheet.title)

# 取得活頁簿的「作用工作表」
anotherSheet = wb.active
print(anotherSheet)

#========================================

# 取得儲存格的內含值
cell_value = sheet['A1'].value
print(cell_value)

# 取得儲存格（物件）
rng = sheet['B2']
print(rng)

# 取得儲存格物件的列屬性
print(rng.row)

# 取得儲存格物件的欄屬性
print(rng.column)

# 取得儲存格物件的座標屬性
print(rng.coordinate)

#========================================

sheet = wb.get_sheet_by_name('工作表1')

# 使用（列，欄）座標，取出儲存格的內容
sheet.cell(row=1, column=4).value

for i in range(1, 7):
     print(i, sheet.cell(row=1, column=i).value)

for i in range(1, 7):
     print(i, sheet.cell(row=2, column=i).value)
