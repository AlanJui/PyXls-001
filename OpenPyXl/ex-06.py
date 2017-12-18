from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# 載入 Excel 檔案（即：活頁簿），取得所有工作表的名稱
wb = load_workbook('../data/sample.xlsx')

# 依「工作表的名稱」取得工作表
sheet = wb.get_sheet_by_name('工作表1')

#========================================

col_name = get_column_letter(1)
print(col_name)

row_name = column_index_from_string('A')
print(row_name)

#========================================

for cellObj in sheet['A1':'F7']:
     for cell in cellObj:
          print(cell.coordinate, cell.value)
     print('--- EOL ---')

print(sheet.max_row)
print(sheet.max_column)

