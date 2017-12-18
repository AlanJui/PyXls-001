import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Load Excel file data into DataFrame
file = 'sample.xlsx'
dir_path = '../data/'
file_path = dir_path + file
# print(file_path)

# 載入 Excel 檔案（即：活頁簿）
wb = load_workbook(file_path)

# 依「工作表的名稱」取得工作表
sheet = wb.get_sheet_by_name('工作表1')

# convert sheet to DataFrame
df = pd.DataFrame(sheet.values)
print('DataFrame = {}'.format(df))
